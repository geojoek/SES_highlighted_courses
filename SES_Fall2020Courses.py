# Script for reading in excel spreadsheet of Highlighted Fall Courses for U-Mas SES website and dumping content out into formatted HTML file
# by Joe Kopera, April 2020

# *** Note: this presumes the Excel sheet is already sorted by department and course number.  Didn't have time to write subsection of script to sort sheet once read-in and then write out to new excel file as openpyxl doesn't sort natively.
# Could load openpyxl data into dictionary, sort it, and do it that way, but under a deadline and have EBV brain fog.

# As it stands this script just appends to the HTML file... it does not overwrite it.  I need to add that functionality.

workingFolder = r"[FOLDERPATH]" # Where the excel file is stored
excelFile = r"[FILEPATH]" # The name of the excel  file
outFileName = r"[OUTFILEPATH]" # The name of the HTML file to write out to

# Department equivalents based on field contents in spreadsheet that we already know are there:

titleBCT = "Building and Construction Technology"
titleECO = "Environmental Conservation"
titleENVIRSCI = "Environmental Science"
titleGeograph = "Geography"
titleGeol = "Geology"
titleGeosci = "Geosciences"
titleLANDARCH = "Landscape Architecture"
titleMicrobio = "Microbiology"
titleNRC = "Natural Resources Conservation"
titleREGIONPL = "Regonal Planning"
titleSTOCKSCH = "Stockbridge School of Agriculture"
titleSUSTCOMM = "Sustainable Communities"

deptList = [titleBCT, titleECO, titleENVIRSCI, titleGeograph, titleGeol, titleGeosci, titleLANDARCH, titleMicrobio, titleNRC, titleREGIONPL, titleSTOCKSCH, titleSUSTCOMM]

# Import `os`
import os

# Change directory
os.chdir(workingFolder)
print("Changing to working folder %s\n"%(workingFolder))

import openpyxl
from openpyxl import load_workbook

workBook = load_workbook(excelFile) # loads Excel File
print("Loading %s\n"%(excelFile))

print("The sheets in this workbook are {} \n".format(workBook.sheetnames))

sheet = workBook.active # selects first available sheet in excel workbook file

def getColumnNames(inputSheet): # function that prints out column names as check to see if openpyxl is parsing correctly
    for column in inputSheet[1]:
        if column.value != None:
            print(column.value)
        else:
            pass

print("The names of columns in this sheet are:\n")
getColumnNames(sheet)

print("Iterating through rows and parsing sheet into HTML file: {}".format(outFileName))

# Begins parser section of script.

with open(outFileName, 'a') as outFile: # creates and opens HTML file for writing. Putting it here so script doesn't open and close this file a thousand times

    outFile.write("<h4>\n")
    for x in deptList:
        outFile.write("| <a href=\"#{}\">{}</a> ".format(x, x))
    outFile.write("\n</h4>")

    index = 0
    for row in sheet.iter_rows(): # iterates through rows in sheet and assigns index to each row since I can't find whatever the native method or class for that in openpyxl

    # setting up indexer so that script can check value of previous row in order to insert department name when it changes in Excel file.  Assumes excel file is sorted correctly.

        index = index + 1

        if row == sheet[1]: # skips first row since it is column headers
            pass

        else:

            prevRowIndex = "A" + str(index - 1)
            prevCourseDept = sheet[prevRowIndex].value

            if row[0].value == None: # skips blank rows if value of first column in row is Null
                pass

            else: # Actual parsing starts now. Based on Spreadsheet having pre-ordained structure

                # retrieving cell values from spreadsheet based on their index location of tuple returned by sheet.iter_rows
                courseDept = row[0].value
                courseNumber = row[1].value
                courseTitle = row[2].value
                courseProf = row[3].value
                courseDesc = row[4].value

                # parsing cell values into html
                htmlParsedCourse = "\n<p>" + "<strong>" + str(courseNumber) + " " + courseTitle + " — " + courseProf + "</strong><br>" + courseDesc + "\n</p>\n"

                if courseDept == prevCourseDept:
                    outFile.write(htmlParsedCourse) # writes out html to file

                else: # subroutine that inserts department titles when they change
                    outFile.write("\n")
                    if courseDept == "BCT":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleBCT, titleBCT))
                    elif courseDept == "ENVIRSCI":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleENVIRSCI, titleENVIRSCI ))
                    elif courseDept == "ECO":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleECO, titleECO))
                    elif courseDept == "Geograph":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleGeograph, titleGeograph))
                    elif courseDept == "Geol":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleGeol, titleGeol))
                    elif courseDept == "Geosci":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleGeosci, titleGeosci))
                    elif courseDept == "LANDARCH":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleLANDARCH, titleLANDARCH))
                    elif courseDept == "Microbio":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleMicrobio, titleMicrobio))
                    elif courseDept == "NRC":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleNRC, titleNRC))
                    elif courseDept == "REGIONPL":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleREGIONPL, titleREGIONPL))
                    elif courseDept == "STOCKSCH":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleSTOCKSCH, titleSTOCKSCH))
                    elif courseDept == "SUSTCOMM":
                        outFile.write("<h2><a id=\"{}\"></a>{}</h2>\n".format(titleSUSTCOMM, titleSUSTCOMM))
                    outFile.write(htmlParsedCourse)